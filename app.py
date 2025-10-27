# app.py
# -----------------------------------------------
# Flask app: routes, uploads, data APIs, downloads, printing.
# Includes careful error handling so the frontend never sees raw HTML on fetch.
# -----------------------------------------------

from __future__ import annotations
import os, io, base64, tempfile, platform, subprocess, traceback, db
from typing import Any, List, Dict
from flask import Flask, request, jsonify, send_file, render_template
import fitz

from db import (
    init_db, get_all, upsert_orders, assign_product_id,
    assign_barcode, confirm_extra, reload_masters, DB_PATH
)
from parser import parse_labels_from_pdf

app = Flask(__name__, template_folder="templates", static_folder="static")
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Optional Windows helper for silent printing
PDFTOPRINTER = os.path.join(BASE_DIR, "bin", "PDFtoPrinter.exe")

# ---------- Utilities ----------

def _json_error(msg: str, status: int = 400):
    """Always return JSON on API errors (prevents HTML dump in the UI)."""
    return jsonify({"ok": False, "error": msg}), status

def _safe_int(s, default=0):
    try:
        return int(s)
    except Exception:
        return default

# ---------- Routes ----------
@app.route("/pending_skus/<contact>")
def pending_skus(contact):
    rows = db.get_all()
    contact = contact.strip()
    pending = [r for r in rows if r["contact_number"] == contact and db._row_remaining_units(r) > 0]
    pending_skus = list({r["sku"] for r in pending if r["sku_type"] != "NoScan"})
    return jsonify({"pending_skus": pending_skus})

@app.route("/sku_contact/<sku>")
def sku_contact(sku):
    """
    Returns the contact_number linked to the next unassigned unit of the given SKU.
    Helps frontend decide if a scan belongs to current locked order.
    """
    sku = sku.strip().upper()
    rows = db.get_all()
    for r in rows:
        if db._normalize_sku_for_db(r.get("sku", "")) == sku and db._row_remaining_units(r) > 0:
            return jsonify({"contact_number": r["contact_number"]})
    return jsonify({"contact_number": None})


@app.get("/")
def home():
    return render_template("index.html")

@app.get("/data")
def data():
    try:
        rows = get_all()
        return jsonify(rows)
    except Exception as e:
        app.logger.exception("DATA_FETCH_FAILED")
        return _json_error(f"Data fetch failed: {e}", 500)

@app.post("/upload")
def upload():
    """
    Upload a PDF, parse per-page rows, upsert into CSV.
    We store the uploaded file as uploads_<original>.pdf in project root.
    """
    try:
        if "file" not in request.files:
            return _json_error("No file part", 400)
        f = request.files["file"]
        if not f or not f.filename.lower().endswith(".pdf"):
            return _json_error("Please upload a PDF", 400)

        safe_name = "uploads_" + os.path.basename(f.filename).replace(" ", "_")
        save_path = os.path.join(BASE_DIR, safe_name)
        f.save(save_path)

        parsed = parse_labels_from_pdf(save_path)
        # annotate rows with source_file
        for r in parsed:
            r["source_file"] = safe_name

        added = upsert_orders(parsed, source_file=safe_name)
        return jsonify({"ok": True, "parsed": len(parsed), "added": added})
    except Exception as e:
        app.logger.exception("UPLOAD_FAILED")
        return _json_error(f"Upload failed: {e}", 500)

@app.post("/assign")
def assign_legacy():
    """
    Legacy: set the same product_id to all rows of a given AWB.
    """
    try:
        payload = request.get_json(force=True, silent=False)
        awb = (payload or {}).get("awb", "").strip()
        product_id = (payload or {}).get("product_id", "").strip().upper()
        ok, msg, status = assign_product_id(awb, product_id)
        if not ok:
            return _json_error(msg, status)
        return jsonify({"ok": True, "message": msg})
    except Exception as e:
        app.logger.exception("ASSIGN_FAILED")
        return _json_error(f"Assign failed: {e}", 500)

@app.post("/assign_barcode")
def assign_barcode_route():
    """
    Modern flow: parse barcode, find matching row (by SKU, qty remaining),
    append token, return progress & print_info upon group completion.
    """
    try:
        payload = request.get_json(force=True, silent=False)
        code = (payload or {}).get("barcode", "").strip()
        active_awb = (payload or {}).get("awb", None)
        ok, data_or_msg, status = assign_barcode(code, active_awb)
        if not ok:
            return _json_error(data_or_msg, status)
        return jsonify({"ok": True, **(data_or_msg if isinstance(data_or_msg, dict) else {})})
    except Exception as e:
        app.logger.exception("ASSIGN_BARCODE_FAILED")
        return _json_error(f"Barcode assign failed: {e}", 500)

@app.post("/confirm_extra")
def confirm_extra_route():
    try:
        payload = request.get_json(force=True, silent=False)
        row_id = (payload or {}).get("row_id", "").strip()
        ok, msg, status = confirm_extra(row_id)
        if not ok:
            return _json_error(msg, status)
        return jsonify({"ok": True, "message": msg})
    except Exception as e:
        app.logger.exception("CONFIRM_EXTRA_FAILED")
        return _json_error(f"Confirm failed: {e}", 500)

@app.get("/download_csv")
def download_csv():
    try:
        if not os.path.exists(DB_PATH):
            return _json_error("No data yet", 404)
        return send_file(
            DB_PATH,
            mimetype="text/csv",
            as_attachment=True,
            download_name="orders_db.csv"
        )
    except Exception as e:
        app.logger.exception("CSV_DOWNLOAD_FAILED")
        return _json_error(f"CSV download failed: {e}", 500)

@app.get("/download_xlsx")
def download_xlsx():
    """
    Excel export (requires openpyxl).
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter

        rows = get_all()
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = ["order_date","order_id","customer_name","contact_number","product_name","sku",
                   "quantity","awb","product_id","row_id","created_at","source_file","page_index","sku_type"]
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h, "") for h in headers])
        # autosize
        for col_idx, _ in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

        bio = io.BytesIO()
        wb.save(bio); bio.seek(0)
        return send_file(
            bio, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True, download_name="orders_db.xlsx"
        )
    except ModuleNotFoundError:
        return _json_error("openpyxl not installed. pip install openpyxl", 500)
    except Exception as e:
        app.logger.exception("XLSX_DOWNLOAD_FAILED")
        return _json_error(f"XLSX download failed: {e}", 500)

@app.get("/label_print")
def label_print():
    """
    Render a single PDF page as PNG inside an HTML that auto-calls window.print().
    Query: ?source=<uploads_*.pdf>&page=<1-based int>
    """
    try:
        src = (request.args.get("source") or "").strip()
        page = (request.args.get("page") or "").strip()

        if not src or not src.lower().endswith(".pdf") or not src.startswith("uploads_"):
            return _json_error("Invalid source", 400)
        page_i = _safe_int(page, 0)
        if page_i < 1:
            return _json_error("Invalid page", 400)

        pdf_path = os.path.join(BASE_DIR, src)
        if not os.path.exists(pdf_path):
            return _json_error("File not found", 404)

        doc = fitz.open(pdf_path)
        if page_i > len(doc):
            doc.close()
            return _json_error("Page out of range", 404)
        pg = doc.load_page(page_i - 1)
        pix = pg.get_pixmap(matrix=fitz.Matrix(3, 3), alpha=False)
        img_b = pix.tobytes("png")
        doc.close()

        b64 = base64.b64encode(img_b).decode("ascii")
        html = f"""<!doctype html>
<html><head><meta charset="utf-8"><title>Print Label</title>
<style>html,body{{margin:0;background:#fff}} img{{width:100%;height:auto;display:block}}</style>
</head><body><img src="data:image/png;base64,{b64}" alt="label"/>
<script>window.onload=()=>{{try{{window.print()}}catch(_){{
}}}};</script></body></html>"""
        return html, 200, {"Content-Type":"text/html; charset=utf-8"}
    except Exception as e:
        app.logger.exception("LABEL_PRINT_FAILED")
        return _json_error(f"Render failed: {e}", 500)

@app.post("/label_print_silent")
def label_print_silent():
    """
    Body: { "source": "uploads_*.pdf", "page": 1, "printer": "Name of Printer" }
    Windows: uses PDFtoPrinter.exe /s to print silently to the specific printer.
    macOS/Linux fallback: uses `lp -d <printer> file.pdf`.
    """
    try:
        data = request.get_json(force=True)
        src = (data or {}).get("source", "").strip()
        page = (data or {}).get("page", "")
        printer = (data or {}).get("printer", "").strip()

        if not (src and str(page).isdigit() and int(page) >= 1 and printer):
            return _json_error("source, page (>=1), printer are required", 400)
        if not src.lower().endswith(".pdf") or not src.startswith("uploads_"):
            return _json_error("Invalid source", 400)

        pdf_path = os.path.join(BASE_DIR, src)
        if not os.path.exists(pdf_path):
            return _json_error("Source file not found", 404)

        page_i = int(page) - 1
        doc = fitz.open(pdf_path)
        if page_i >= len(doc):
            doc.close()
            return _json_error("Page out of range", 404)
        one = fitz.open()
        one.insert_pdf(doc, from_page=page_i, to_page=page_i)
        tmpdir = tempfile.mkdtemp(prefix="print_")
        one_pdf = os.path.join(tmpdir, "label.pdf")
        one.save(one_pdf); one.close(); doc.close()

        if platform.system() == "Windows":
            if not os.path.exists(PDFTOPRINTER):
                return _json_error("PDFtoPrinter.exe not found in ./bin", 500)
            cmd = [PDFTOPRINTER, one_pdf, printer, "/s"]
        else:
            cmd = ["lp", "-d", printer, one_pdf]

        res = subprocess.run(cmd, capture_output=True, text=True)
        if res.returncode != 0:
            return _json_error(f"Print failed: {res.stderr or res.stdout or res.returncode}", 500)

        return jsonify({"ok": True, "message": f"Sent to printer '{printer}'"})
    except Exception as e:
        app.logger.exception("LABEL_PRINT_SILENT_FAILED")
        return _json_error(f"Render/print failed: {e}", 500)

@app.post("/admin/reload_masters")
def admin_reload():
    try:
        reload_masters()
        return jsonify({"ok": True})
    except Exception as e:
        app.logger.exception("RELOAD_FAILED")
        return _json_error(f"Reload failed: {e}", 500)

# ---------- App start ----------
if __name__ == "__main__":
    try:
        init_db()
        app.run(host="127.0.0.1", port=5000, debug=False)
    except Exception:
        traceback.print_exc()
